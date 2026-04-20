import os
import io
import json
import sqlite3
from contextlib import contextmanager
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import streamlit as st

try:
    from google import genai
except Exception:
    genai = None

try:
    import cv2
except Exception:
    cv2 = None

try:
    import librosa
except Exception:
    librosa = None

try:
    import numpy as np
except Exception:
    np = None

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    REPORTLAB_OK = True
except Exception:
    REPORTLAB_OK = False

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    EXCEL_OK = True
except Exception:
    EXCEL_OK = False

APP_NAME = "LCA_PRO"
BASE_DIR = Path("lca_pro_data")
DB_PATH = BASE_DIR / "lca_pro.db"
UPLOAD_DIR = BASE_DIR / "uploads"
BASE_DIR.mkdir(parents=True, exist_ok=True)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

AREAS = ["supervision_tecnica", "orden_trabajo", "seguridad_sst", "logistica_compras"]
ESTADOS = ["Pendiente", "Aprobado", "Desaprobado"]
AREA_LABELS = {
    "supervision_tecnica": "Supervisión Técnica",
    "orden_trabajo": "Orden de Trabajo",
    "seguridad_sst": "Seguridad / SST",
    "logistica_compras": "Logística / Compras",
}
PDF_TITLES = {
    "supervision_tecnica": "Informe Técnico Especializado",
    "orden_trabajo": "Orden de Trabajo",
    "seguridad_sst": "Matriz de Seguridad SST",
    "logistica_compras": "Presupuesto y Requerimiento de Compra",
}


def is_technician():
    return st.session_state.get("usuario_rol") == "tecnico"


def is_report_viewer():
    return st.session_state.get("usuario_rol") in ["admin", "usuario"]


def company_restricted_query(query, params=()):
    empresa_id = st.session_state.get("empresa_id")
    if empresa_id is None:
        return []
    if "WHERE" in query.upper():
        return q(query + " AND c.empresa_id = ?", tuple(params) + (empresa_id,), fetch=True)
    return q(query + " WHERE empresa_id = ?", tuple(params) + (empresa_id,), fetch=True)

EXPERT_PROMPTS = {
    "manguera_hidraulica": {
        "system": "Actúa como perito senior de maquinaria pesada especializado en sistemas hidráulicos. Evalúa daño real del componente visible, priorizando seguridad y operatividad. No aceptes respuestas preventivas genéricas si existe deterioro visible.",
        "focus": ["cubierta externa", "malla expuesta", "abrasión", "cortes", "cuarteamiento", "deformación", "fuga", "contaminación", "riesgo de estallido", "terminales"],
    },
    "hidraulico": {
        "system": "Actúa como perito senior en hidráulica de equipo pesado. Evalúa mangueras, cilindros, bomba, válvulas, manifold, tanque y conexiones. Prioriza fugas, contaminación, presión, integridad física y criticidad operacional.",
        "focus": ["fuga", "presión", "mangueras", "terminales", "cilindros", "sellos", "abrasión", "temperatura", "contaminación", "cavitación"],
    },
    "motor": {
        "system": "Actúa como perito senior de motores diésel de equipo pesado. Evalúa condición real del sistema y evita teoría general.",
        "focus": ["fuga de aceite", "humo", "ruido metálico", "sobrecalentamiento", "inyectores", "turbo", "correas", "vibración", "compresión"],
    },
    "neumatico": {
        "system": "Actúa como perito senior en neumáticos de flota y maquinaria pesada. Evalúa desgaste real, hombros, costados, cortes, deformaciones y seguridad operacional.",
        "focus": ["hombro interno", "hombro externo", "banda de rodamiento", "costado", "corte", "abultamiento", "presión", "alineación", "sobrecarga", "fatiga"],
    },
    "transmision": {
        "system": "Actúa como perito senior en transmisión de equipo pesado. Evalúa caja, cardán, crucetas, diferencial y soportes evitando respuestas genéricas.",
        "focus": ["holgura", "ruido", "vibración", "fuga", "engranajes", "cardán", "crucetas", "rodamientos", "soportes", "alineación"],
    },
    "frenos": {
        "system": "Actúa como perito senior en sistemas de freno de maquinaria y flota pesada. Evalúa seguridad, criticidad, desgaste y condición operativa.",
        "focus": ["pastillas", "discos", "caliper", "fuga", "temperatura", "desgaste", "fisura", "roce", "líneas", "pérdida de eficiencia"],
    },
}
COMPONENT_ALIAS = {
    "manguera": "manguera_hidraulica",
    "manguera hidraulica": "manguera_hidraulica",
    "hidraulico": "hidraulico",
    "motor": "motor",
    "neumatico": "neumatico",
    "llanta": "neumatico",
    "transmision": "transmision",
    "caja": "transmision",
    "diferencial": "transmision",
    "frenos": "frenos",
}

def inject_css():
    st.markdown("""
    <style>
    .stApp { background: linear-gradient(180deg, #f8fafc 0%, #eef2ff 100%); }
    .block-container { padding-top: 1.1rem; padding-bottom: 2rem; max-width: 1450px; }
    h1, h2, h3 { color: #0f172a !important; font-weight: 700 !important; }
    .card { background: white; border-radius: 18px; padding: 18px 18px 10px 18px; box-shadow: 0 8px 24px rgba(15, 23, 42, 0.08); border: 1px solid #e2e8f0; margin-bottom: 18px; }
    .hero { background: linear-gradient(135deg, #0f172a 0%, #1e40af 55%, #2563eb 100%); color: white; padding: 22px 24px; border-radius: 20px; box-shadow: 0 10px 28px rgba(30, 64, 175, 0.24); margin-bottom: 18px; }
    .hero h1, .hero p { color: white !important; margin: 0 !important; }
    .section-title { font-weight: 800; color: #0f172a; margin-bottom: 8px; }
    .status-badge { display: inline-block; padding: 8px 14px; border-radius: 999px; font-weight: 700; font-size: 0.9rem; margin-top: 4px; margin-bottom: 12px; }
    .status-pendiente {background: #dbeafe; color:#1d4ed8;}
    .status-parcial {background: #fef3c7; color:#b45309;}
    .status-observacion {background: #fee2e2; color:#b91c1c;}
    .status-final {background: #dcfce7; color:#166534;}
    .kpi { border-radius: 18px; padding: 16px; color: white; min-height: 96px; box-shadow: 0 8px 20px rgba(15, 23, 42, 0.10); }
    .kpi-title { font-size: 0.95rem; opacity: .9; margin-bottom: 8px; }
    .kpi-value { font-size: 1.2rem; font-weight: 800; }
    .kpi-blue {background: linear-gradient(135deg, #1e3a8a, #2563eb);}
    .kpi-green {background: linear-gradient(135deg, #166534, #16a34a);}
    .kpi-amber {background: linear-gradient(135deg, #b45309, #f59e0b);}
    .kpi-red {background: linear-gradient(135deg, #991b1b, #dc2626);}
    .small-note { color: #475569; font-size: 0.92rem; }
    div.stButton > button, div.stDownloadButton > button { border-radius: 12px !important; font-weight: 700 !important; }
    @media (max-width: 768px) {
        .block-container { padding-top: 0.8rem !important; padding-bottom: 1.2rem !important; max-width: 100% !important; }
        .stMarkdown, .stText, .stTable, .stDataFrame { font-size: 0.95rem !important; }
        .hero { padding: 18px 16px; }
        .card { padding: 14px 14px 10px 14px; }
        .kpi { min-height: 80px; padding: 14px; }
        .kpi-value { font-size: 1.05rem; }
        .css-1lcbmhc.e1fqkh3o3, .css-1d391kg { width: 100% !important; }
        .element-container .stButton { width: 100% !important; }
    }
    </style>
    """, unsafe_allow_html=True)

@contextmanager
def db_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()

def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def q(query, params=(), fetch=False, one=False):
    with db_conn() as conn:
        cur = conn.execute(query, params)
        if fetch:
            rows = cur.fetchall()
            data = [dict(r) for r in rows]
            return data[0] if one and data else (None if one else data)
        return cur.lastrowid

def init_db():
    with db_conn() as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS empresas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT UNIQUE NOT NULL,
            descripcion TEXT,
            email TEXT,
            telefono TEXT,
            direccion TEXT,
            ciudad TEXT,
            pais TEXT,
            activa INTEGER DEFAULT 1,
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario TEXT NOT NULL,
            password TEXT NOT NULL,
            nombre_completo TEXT,
            email TEXT,
            telefono TEXT,
            empresa_id INTEGER NOT NULL,
            rol TEXT DEFAULT 'usuario',
            activo INTEGER DEFAULT 1,
            created_at TEXT,
            UNIQUE(usuario, empresa_id),
            FOREIGN KEY(empresa_id) REFERENCES empresas(id)
        );
        CREATE TABLE IF NOT EXISTS cases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            empresa_id INTEGER NOT NULL,
            unidad TEXT, placa TEXT, componente TEXT, motivo TEXT, horas INTEGER, observaciones TEXT,
            diagnostico_base TEXT, image_path TEXT, audio_path TEXT, video_path TEXT,
            sensor_tipo TEXT, sensor_valor TEXT, sensor_unidad TEXT, sensor_min TEXT, sensor_max TEXT,
            sensor_obs TEXT, ai_model TEXT, prompt_key TEXT, created_at TEXT,
            FOREIGN KEY(empresa_id) REFERENCES empresas(id)
        );
        CREATE TABLE IF NOT EXISTS area_docs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            case_id INTEGER NOT NULL, area TEXT NOT NULL, contenido_json TEXT NOT NULL, updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS approvals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            case_id INTEGER NOT NULL, area TEXT NOT NULL, estado TEXT NOT NULL, responsable TEXT, comentario TEXT, fecha TEXT NOT NULL
        );
        """)


def table_has_column(table_name: str, column_name: str) -> bool:
    columns = q(f"PRAGMA table_info({table_name})", fetch=True)
    return any(col.get("name") == column_name for col in columns)


def migrate_schema():
    # Ensure the cases table has empresa_id for older existing databases.
    if not table_has_column("cases", "empresa_id"):
        default_empresa = q("SELECT id FROM empresas ORDER BY id LIMIT 1", fetch=True, one=True)
        if not default_empresa:
            default_empresa_id = q(
                "INSERT INTO empresas(nombre, descripcion, activa, created_at) VALUES(?,?,?,?)",
                ("EQUIPRO Perú", "Empresa por defecto para datos heredados", 1, now_str())
            )
        else:
            default_empresa_id = default_empresa["id"]
        q("ALTER TABLE cases ADD COLUMN empresa_id INTEGER")
        q("UPDATE cases SET empresa_id = ? WHERE empresa_id IS NULL", (default_empresa_id,))

    # Add new columns to empresas
    if not table_has_column("empresas", "email"):
        q("ALTER TABLE empresas ADD COLUMN email TEXT")
    if not table_has_column("empresas", "telefono"):
        q("ALTER TABLE empresas ADD COLUMN telefono TEXT")
    if not table_has_column("empresas", "direccion"):
        q("ALTER TABLE empresas ADD COLUMN direccion TEXT")
    if not table_has_column("empresas", "ciudad"):
        q("ALTER TABLE empresas ADD COLUMN ciudad TEXT")
    if not table_has_column("empresas", "pais"):
        q("ALTER TABLE empresas ADD COLUMN pais TEXT")

    # Add new columns to usuarios
    if not table_has_column("usuarios", "nombre_completo"):
        q("ALTER TABLE usuarios ADD COLUMN nombre_completo TEXT")
    if not table_has_column("usuarios", "email"):
        q("ALTER TABLE usuarios ADD COLUMN email TEXT")
    if not table_has_column("usuarios", "telefono"):
        q("ALTER TABLE usuarios ADD COLUMN telefono TEXT")


def init_demo_data():
    """Carga datos de demostración de empresas y usuarios"""
    try:
        # Verificar si ya existen empresas
        existing = q("SELECT COUNT(*) as cnt FROM empresas", fetch=True, one=True)
        if existing and existing['cnt'] > 0:
            return
        
        # Agregar empresas de demostración
        empresas = [
            ("EQUIPRO Perú", "Empresa de servicios técnicos", "contacto@equipro.pe", "+51 123 456 789", "Av. Industrial 123", "Lima", "Perú", 1, now_str()),
            ("MINERÍA ANDINA", "Servicios de mantenimiento minería", "info@mineriaandina.com", "+51 987 654 321", "Calle Minera 456", "Arequipa", "Perú", 1, now_str()),
            ("LOGÍSTICA EXPRESS", "Flota y mantenimiento", "ventas@logisticaexpress.pe", "+51 555 123 456", "Jr. Transporte 789", "Cusco", "Perú", 1, now_str()),
        ]
        
        empresa_ids = {}
        for nombre, desc, email, tel, dir, ciudad, pais, activa, fecha in empresas:
            try:
                empresa_id = q(
                    "INSERT INTO empresas(nombre, descripcion, email, telefono, direccion, ciudad, pais, activa, created_at) VALUES(?,?,?,?,?,?,?,?,?)",
                    (nombre, desc, email, tel, dir, ciudad, pais, activa, fecha)
                )
                empresa_ids[nombre] = empresa_id
            except:
                # Si la empresa ya existe, obtener su ID
                result = q("SELECT id FROM empresas WHERE nombre = ?", (nombre,), fetch=True, one=True)
                if result:
                    empresa_ids[nombre] = result['id']
        
        # Agregar usuarios de demostración para cada empresa
        usuarios_demo = [
            ("usuario1", "123456", "Juan Pérez", "juan.perez@equipro.pe", "+51 111 222 333", "EQUIPRO Perú", "admin"),
            ("usuario2", "123456", "María García", "maria.garcia@mineriaandina.com", "+51 444 555 666", "MINERÍA ANDINA", "admin"),
            ("usuario3", "123456", "Carlos López", "carlos.lopez@logisticaexpress.pe", "+51 777 888 999", "LOGÍSTICA EXPRESS", "admin"),
            ("tecnico1", "123456", "Pedro Ramírez", "pedro.ramirez@equipro.pe", "+51 000 111 222", "EQUIPRO Perú", "tecnico"),
            ("tecnico2", "123456", "Ana Torres", "ana.torres@mineriaandina.com", "+51 333 444 555", "MINERÍA ANDINA", "tecnico"),
        ]
        
        for user, password, nombre_completo, email, tel, empresa_nombre, rol in usuarios_demo:
            empresa_id = empresa_ids.get(empresa_nombre)
            if empresa_id:
                try:
                    q(
                        "INSERT INTO usuarios(usuario, password, nombre_completo, email, telefono, empresa_id, rol, activo, created_at) VALUES(?,?,?,?,?,?,?,?,?)",
                        (user, password, nombre_completo, email, tel, empresa_id, rol, 1, now_str())
                    )
                except:
                    pass  # Usuario ya existe
    except Exception as e:
        print(f"Error en init_demo_data: {e}")

def save_upload(uploaded_file, prefix):
    if uploaded_file is None:
        return ""
    safe_name = f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uploaded_file.name}"
    path = UPLOAD_DIR / safe_name
    with open(path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return str(path)

def analyze_image(path: str):
    if not path:
        return {"status": "no_image"}
    if cv2 is None:
        return {"status": "opencv_unavailable"}
    try:
        img = cv2.imread(path)
        if img is None:
            return {"status": "unreadable"}
        h, w = img.shape[:2]
        return {"status": "ok", "width": w, "height": h, "note": "Imagen recibida."}
    except Exception as e:
        return {"status": "error", "detail": str(e)}

def analyze_audio(path: str):
    if not path:
        return {"status": "no_audio"}
    if librosa is None or np is None:
        return {"status": "audio_libs_unavailable"}
    try:
        y, sr = librosa.load(path, sr=None)
        duration = len(y) / sr if sr else 0
        peak = float(np.max(np.abs(y))) if len(y) else 0
        return {"status": "ok", "sample_rate": sr, "duration_sec": round(duration, 2), "peak_amplitude": round(peak, 4), "note": "Audio recibido."}
    except Exception as e:
        return {"status": "error", "detail": str(e)}

def analyze_video(path: str):
    if not path:
        return {"status": "no_video"}
    if cv2 is None:
        return {"status": "opencv_unavailable"}
    try:
        cap = cv2.VideoCapture(path)
        if not cap.isOpened():
            return {"status": "unreadable"}
        frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        fps = float(cap.get(cv2.CAP_PROP_FPS))
        duration = round(frames / fps, 2) if fps else 0
        cap.release()
        return {"status": "ok", "frames": frames, "fps": round(fps, 2), "duration_sec": duration, "note": "Video recibido."}
    except Exception as e:
        return {"status": "error", "detail": str(e)}

def get_models():
    preferred = os.getenv("LCA_EQUIPRO_MODEL", "").strip()
    ordered = [preferred, "gemini-2.5-flash", "gemini-2.5-pro", "gemini-3-flash-preview"]
    out = []
    seen = set()
    for m in ordered:
        if m and m not in seen:
            out.append(m)
            seen.add(m)
    return out

def call_ai(system_prompt: str, user_prompt: str):
    if genai is None:
        return None, None, ["google-genai no está instalado"]
    api_key = os.getenv("GEMINI_API_KEY", "").strip()
    if not api_key:
        try:
            api_key = st.secrets.get("GEMINI_API_KEY", "").strip()
        except Exception:
            pass
    if not api_key:
        return None, None, ["GEMINI_API_KEY no configurada"]
    client = genai.Client(api_key=api_key)
    composed = f"{system_prompt}\n\n{user_prompt}"
    errors = []
    for model in get_models():
        try:
            resp = client.models.generate_content(model=model, contents=composed)
            txt = (resp.text or "").strip()
            if txt:
                return txt, model, errors
        except Exception as e:
            errors.append(f"{model}: {e}")
    return None, None, errors

def normalize_text(text: str) -> str:
    if not text:
        return ""
    return text.translate(str.maketrans("áéíóúÁÉÍÓÚüÜñÑ", "aeiouAEIOUuUNN")).strip().lower()


def get_prompt_key(component: str, observations: str):
    c = normalize_text(component or "")
    o = normalize_text(observations or "")
    if c in COMPONENT_ALIAS:
        return COMPONENT_ALIAS[c]
    if "manguera" in c or "manguera" in o:
        return "manguera_hidraulica"
    if "hidraul" in c or "hidraul" in o or "valvula" in c or "valvula" in o or "bomba" in c or "bomba" in o or "cilindro" in c or "cilindro" in o or "manifold" in c or "manifold" in o:
        return "hidraulico"
    if "motor" in c or "motor" in o or "diesel" in c or "diesel" in o or "inyect" in c or "inyect" in o or "turbo" in c or "turbo" in o:
        return "motor"
    if "freno" in c or "freno" in o or "disco" in c or "disco" in o or "caliper" in c or "caliper" in o or "pastilla" in c or "pastilla" in o:
        return "frenos"
    if "llanta" in c or "llanta" in o or "neumatico" in c or "neumatico" in o or "rodado" in c or "rodado" in o:
        return "neumatico"
    if "transmision" in c or "transmision" in o or "caja" in c or "caja" in o or "diferencial" in c or "diferencial" in o or "cardan" in c or "cardan" in o or "embrague" in c or "embrague" in o:
        return "transmision"
    if c in ("falla", "fallo", "componente", "problema", "reporte", "servicio", "inspeccion", "revision", "correctivo", "preventivo", "mantenimiento"):
        if "manguera" in o:
            return "manguera_hidraulica"
        if "hidraul" in o:
            return "hidraulico"
        if "motor" in o:
            return "motor"
        if "freno" in o or "disco" in o or "caliper" in o:
            return "frenos"
        if "llanta" in o or "neumatico" in o or "rodado" in o:
            return "neumatico"
        if "transmision" in o or "caja" in o or "diferencial" in o or "cardan" in o or "embrague" in o:
            return "transmision"
        return "hidraulico"
    return "hidraulico"


def build_expert_prompts(data, image_info, audio_info, video_info):
    prompt_key = get_prompt_key(data["componente"], data["obs"])
    expert = EXPERT_PROMPTS[prompt_key]
    system_prompt = expert["system"]
    user_prompt = f"""
COMPONENTE / ESPECIALIDAD APLICADA: {prompt_key}
FOCOS OBLIGATORIOS: {", ".join(expert["focus"])}

CASO:
Unidad: {data['unidad']}
Placa: {data['placa']}
Componente reportado: {data['componente']}
Motivo: {data['motivo']}
Horas: {data['horas']}
Observaciones técnico: {data['obs']}

EVIDENCIA:
Imagen: {image_info}
Audio: {audio_info}
Video: {video_info}
Sensor tipo: {data['sensor_tipo']}
Sensor valor: {data['sensor_valor']}
Sensor unidad: {data['sensor_unidad']}
Sensor rango mínimo: {data['sensor_min']}
Sensor rango máximo: {data['sensor_max']}
Sensor observación: {data['sensor_obs']}

REGLAS:
- Si la imagen muestra deterioro severo, no respondas preventivo normal.
- Si ves abrasión, exposición de refuerzo, corte, corrosión, cuarteamiento, aplastamiento, deformación o fuga, eleva criticidad.
- El diagnóstico debe ser técnico, específico, detallado y accionable.
- No inventes datos no visibles ni no medidos.
- No des teoría general.
- Enfócate exclusivamente en el componente reportado y en las observaciones del caso. Ignora otros sistemas o partes no mencionadas.
- Si el componente es genérico, usa la evidencia para identificar el sistema principal, pero mantén el análisis en la falla real.
- No hagas análisis de transmisión, motor, frenos u otros sistemas si no están claramente indicados en el caso.
- En cada campo describe condición, mecanismo de falla, impacto operativo y acción sugerida.
- En logística incluye descripciones completas y referencias de mercado estimadas cuando no exista código exacto.
- En seguridad detalla peligro, riesgo, control, ATS y KPI en lenguaje profesional.

Responde exactamente en JSON con esta estructura:
{{
  "diagnostico_base": "Párrafo técnico claro, profesional y no genérico. Debe indicar condición real del componente, nivel de criticidad y recomendación inmediata.",
  "supervision_tecnica": {{
    "hallazgos": "Lista o párrafo detallado de hallazgos visibles y funcionales.",
    "diagnostico": "Diagnóstico profesional detallado del componente y su condición operativa.",
    "criticidad": "Baja/Media/Alta/Crítica",
    "pruebas": "Pruebas de confirmación recomendadas, secuencia técnica y criterios de aceptación/rechazo.",
    "recomendaciones": "Acciones correctivas, preventivas y restricciones operativas."
  }},
  "orden_trabajo": {{
    "actividades": "Secuencia técnica detallada de trabajos a ejecutar.",
    "personal": "Perfil del personal requerido con especialidad.",
    "hh_estimadas": "Horas-hombre estimadas con criterio técnico.",
    "herramientas": "Herramientas, equipos y medios auxiliares requeridos.",
    "repuestos": "Repuestos, kits, sellos, conexiones, fluidos o elementos de sustitución necesarios."
  }},
  "seguridad_sst": {{
    "peligro": "Peligros específicos del trabajo y del componente.",
    "riesgo": "Riesgos operacionales, ambientales y a las personas.",
    "controles": "Controles preventivos y de mitigación concretos.",
    "ats": "Secuencia de ATS detallada, bloqueo, alivio de presión, aislamiento, limpieza y verificación final.",
    "epp": "EPP obligatorio específico por tarea.",
    "kpi": "Indicadores y metas de control relevantes al caso."
  }},
  "logistica_compras": {{
    "piezas": "Detalle técnico-profesional de piezas o kits a comprar, con especificación suficiente para cotizar.",
    "consumibles": "Consumibles, fluidos, limpiadores, paños, selladores u otros materiales requeridos.",
    "mano_obra": "Especialidad y alcance de mano de obra requerida.",
    "costo_piezas": "Monto estimado numérico o referencia de mercado.",
    "costo_mo": "Monto estimado numérico o referencia de mercado.",
    "prioridad": "Normal/Urgente/Programada",
    "justificacion": "Justificación técnica profesional para adquisición y ejecución."
  }}
}}
""".strip()
    return prompt_key, system_prompt, user_prompt

def parse_ai_json(txt: str):
    try:
        return json.loads(txt)
    except Exception:
        start = txt.find("{")
        end = txt.rfind("}")
        if start != -1 and end != -1 and end > start:
            try:
                return json.loads(txt[start:end+1])
            except Exception:
                return None
        return None

def seed_demo():
    if q("SELECT * FROM cases LIMIT 1", fetch=True):
        return
    empresa_id = q("SELECT id FROM empresas ORDER BY id LIMIT 1", fetch=True, one=True)
    if not empresa_id:
        return
    case_id = q(
        """INSERT INTO cases(
            empresa_id, unidad, placa, componente, motivo, horas, observaciones, diagnostico_base,
            image_path, audio_path, video_path, sensor_tipo, sensor_valor, sensor_unidad,
            sensor_min, sensor_max, sensor_obs, ai_model, prompt_key, created_at
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (empresa_id['id'], "EXCAVADORA 01", "EX-001", "Manguera", "Correctivo", 150,
         "Inspección a manguera de componente hidráulico con desgaste visible.",
         "Pendiente de análisis experto.",
         "", "", "", "Presión", "0", "psi", "0", "0", "Sin lectura", "", "", now_str())
    )
    for area in AREAS:
        q("INSERT INTO area_docs(case_id, area, contenido_json, updated_at) VALUES(?,?,?,?)", (case_id, area, "{}", now_str()))
        q("INSERT INTO approvals(case_id, area, estado, responsable, comentario, fecha) VALUES(?,?,?,?,?,?)", (case_id, area, "Pendiente", "", "", now_str()))

def get_empresas():
    """Obtiene lista de empresas activas"""
    return q("SELECT id, nombre FROM empresas WHERE activa = 1 ORDER BY nombre", fetch=True)

def validar_login(usuario: str, password: str, empresa_id: int):
    """Valida credenciales de usuario para una empresa"""
    resultado = q(
        "SELECT id, usuario, rol FROM usuarios WHERE usuario = ? AND password = ? AND empresa_id = ? AND activo = 1",
        (usuario, password, empresa_id),
        fetch=True,
        one=True
    )
    return resultado

def get_empresa_nombre(empresa_id: int):
    """Obtiene nombre de empresa por ID"""
    resultado = q("SELECT nombre FROM empresas WHERE id = ?", (empresa_id,), fetch=True, one=True)
    return resultado['nombre'] if resultado else "Desconocida"

def get_all_empresas():
    return q("SELECT * FROM empresas ORDER BY nombre", fetch=True)

def create_empresa(nombre, descripcion, email, telefono, direccion, ciudad, pais):
    return q(
        "INSERT INTO empresas(nombre, descripcion, email, telefono, direccion, ciudad, pais, activa, created_at) VALUES(?,?,?,?,?,?,?,?,?)",
        (nombre, descripcion, email, telefono, direccion, ciudad, pais, 1, now_str())
    )

def update_empresa(id, nombre, descripcion, email, telefono, direccion, ciudad, pais, activa):
    q(
        "UPDATE empresas SET nombre=?, descripcion=?, email=?, telefono=?, direccion=?, ciudad=?, pais=?, activa=? WHERE id=?",
        (nombre, descripcion, email, telefono, direccion, ciudad, pais, activa, id)
    )

def get_usuarios_by_empresa(empresa_id):
    return q("SELECT * FROM usuarios WHERE empresa_id = ? ORDER BY usuario", (empresa_id,), fetch=True)

def create_usuario(empresa_id, usuario, password, nombre_completo, email, telefono, rol):
    return q(
        "INSERT INTO usuarios(usuario, password, nombre_completo, email, telefono, empresa_id, rol, activo, created_at) VALUES(?,?,?,?,?,?,?,?,?)",
        (usuario, password, nombre_completo, email, telefono, empresa_id, rol, 1, now_str())
    )

def update_usuario(id, usuario, password, nombre_completo, email, telefono, rol, activo):
    q(
        "UPDATE usuarios SET usuario=?, password=?, nombre_completo=?, email=?, telefono=?, rol=?, activo=? WHERE id=?",
        (usuario, password, nombre_completo, email, telefono, rol, activo, id)
    )

def delete_usuario(id):
    q("DELETE FROM usuarios WHERE id=?", (id,))

def get_cases_df():
    """Obtiene casos filtrados por empresa del usuario"""
    empresa_id = st.session_state.get("empresa_id")
    if not empresa_id:
        return pd.DataFrame()
    data = q(
        "SELECT id, placa, unidad, componente, motivo, horas, created_at FROM cases WHERE empresa_id = ? ORDER BY id DESC",
        (empresa_id,),
        fetch=True
    )
    return pd.DataFrame(data) if data else pd.DataFrame()

def get_all_cases_df():
    """Obtiene todos los casos para admin"""
    data = q(
        "SELECT c.id, c.placa, c.unidad, c.componente, c.motivo, c.horas, c.created_at, e.nombre as empresa FROM cases c JOIN empresas e ON c.empresa_id = e.id ORDER BY c.id DESC",
        fetch=True
    )
    return pd.DataFrame(data) if data else pd.DataFrame()

def get_case(case_id):
    """Obtiene caso verificando que pertenezca a la empresa del usuario"""
    empresa_id = st.session_state.get("empresa_id")
    if not empresa_id:
        return None
    return q("SELECT * FROM cases WHERE id = ? AND empresa_id = ?", (case_id, empresa_id), fetch=True, one=True)

def get_doc(case_id, area):
    row = q("SELECT * FROM area_docs WHERE case_id=? AND area=?", (case_id, area), fetch=True, one=True)
    if not row:
        return {}
    try:
        return json.loads(row["contenido_json"])
    except Exception:
        return {}

def save_doc(case_id, area, payload):
    row = q("SELECT * FROM area_docs WHERE case_id=? AND area=?", (case_id, area), fetch=True, one=True)
    data = json.dumps(payload, ensure_ascii=False)
    if row:
        q("UPDATE area_docs SET contenido_json=?, updated_at=? WHERE id=?", (data, now_str(), row["id"]))
    else:
        q("INSERT INTO area_docs(case_id, area, contenido_json, updated_at) VALUES(?,?,?,?)", (case_id, area, data, now_str()))

def get_approval(case_id, area):
    return q("SELECT * FROM approvals WHERE case_id=? AND area=? ORDER BY id DESC LIMIT 1", (case_id, area), fetch=True, one=True)

def save_approval(case_id, area, estado, responsable, comentario):
    q("INSERT INTO approvals(case_id, area, estado, responsable, comentario, fecha) VALUES(?,?,?,?,?,?)", (case_id, area, estado, responsable, comentario, now_str()))

def compute_global_status(case_id):
    statuses = []
    for area in AREAS:
        ap = get_approval(case_id, area)
        statuses.append(ap["estado"] if ap else "Pendiente")
    if any(s == "Desaprobado" for s in statuses):
        return "En Observación"
    if all(s == "Aprobado" for s in statuses):
        return "Aprobado Final"
    if any(s == "Aprobado" for s in statuses):
        return "Aprobado Parcial"
    return "Pendiente"

def render_status_badge(status):
    classes = {
        "Pendiente": "status-pendiente",
        "Aprobado Parcial": "status-parcial",
        "En Observación": "status-observacion",
        "Aprobado Final": "status-final",
    }
    st.markdown(f'<div class="status-badge {classes.get(status, "status-pendiente")}">{status}</div>', unsafe_allow_html=True)

def build_pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="CorpTitle", fontName="Helvetica-Bold", fontSize=18, leading=22, textColor=colors.HexColor("#0f172a"), spaceAfter=10, alignment=TA_LEFT))
    styles.add(ParagraphStyle(name="CorpSub", fontName="Helvetica-Bold", fontSize=11, leading=14, textColor=colors.HexColor("#1e40af"), spaceAfter=6))
    styles.add(ParagraphStyle(name="CorpBody", fontName="Helvetica", fontSize=8.8, leading=11.2, textColor=colors.HexColor("#111827")))
    return styles


def draw_pdf_logo(canvas, page_size):
    width = 80
    height = 24
    x = 1.2 * cm
    y = page_size[1] - 1.35 * cm - height
    canvas.saveState()
    canvas.setFillColor(colors.HexColor("#0f172a"))
    canvas.roundRect(x, y, width, height, 4, fill=True, stroke=False)
    canvas.setFillColor(colors.white)
    canvas.setFont("Helvetica-Bold", 10)
    canvas.drawString(x + 6, y + 12, "LCA_PRO")
    canvas.setFont("Helvetica", 6)
    canvas.drawString(x + 6, y + 4, "Sistema Industrial")
    canvas.restoreState()


def draw_pdf_watermark(canvas, page_size):
    canvas.saveState()
    try:
        watermark_color = colors.Color(0.07, 0.10, 0.17, alpha=0.08)
    except Exception:
        watermark_color = colors.HexColor("#9ca3af")
    canvas.setFillColor(watermark_color)
    canvas.setFont("Helvetica-Bold", 80)
    canvas.translate(page_size[0] / 2, page_size[1] / 2)
    canvas.rotate(45)
    canvas.drawCentredString(0, 0, "LCA_PRO")
    canvas.restoreState()


def build_report_stats(df):
    stats = []
    stats.append({"Métrica": "Total de registros", "Valor": len(df)})
    if "Unidad" in df.columns:
        stats.append({"Métrica": "Unidades únicas", "Valor": int(df["Unidad"].nunique())})
    if "Empresa" in df.columns:
        stats.append({"Métrica": "Empresas únicas", "Valor": int(df["Empresa"].nunique())})
    if "Placa" in df.columns:
        stats.append({"Métrica": "Placas únicas", "Valor": int(df["Placa"].nunique())})
    if "Horas" in df.columns:
        horas = pd.to_numeric(df["Horas"], errors="coerce")
        stats.append({"Métrica": "Horas totales", "Valor": round(float(horas.sum()), 2)})
        stats.append({"Métrica": "Horas promedio", "Valor": round(float(horas.mean()), 2)})
    return pd.DataFrame(stats)


def build_report_pivot_tables(df):
    pivots = {}
    if "Unidad" in df.columns:
        pivots["Por Unidad"] = (
            df.groupby("Unidad").size().reset_index(name="Casos").sort_values(by="Casos", ascending=False)
        )
    if "Empresa" in df.columns:
        pivots["Por Empresa"] = (
            df.groupby("Empresa").size().reset_index(name="Casos").sort_values(by="Casos", ascending=False)
        )
    if "Componente" in df.columns:
        pivots["Por Componente"] = (
            df.groupby("Componente").size().reset_index(name="Casos").sort_values(by="Casos", ascending=False)
        )
    return pivots


def normalize_excel_value(value):
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        try:
            return json.dumps(value, ensure_ascii=False)
        except Exception:
            return str(value)
    return str(value)


def auto_adjust_column_widths(ws, min_width=12, max_width=40):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def style_header(ws):
    header_fill = PatternFill(fill_type="solid", fgColor="0f172a")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="B8B8B8"),
        right=Side(style="thin", color="B8B8B8"),
        top=Side(style="thin", color="B8B8B8"),
        bottom=Side(style="thin", color="B8B8B8"),
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def style_data_rows(ws):
    border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        fill_color = "F8FAFC" if row[0].row % 2 == 0 else None
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill_color:
                cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)


def write_df_to_sheet(ws, df):
    ws.append(list(df.columns))
    style_header(ws)
    for row in df.itertuples(index=False):
        ws.append([normalize_excel_value(value) for value in row])
    style_data_rows(ws)
    ws.freeze_panes = "A2"
    auto_adjust_column_widths(ws)
    ws.sheet_view.zoomScale = 90


def sanitize_sheet_title(title):
    invalid_chars = set('[]:*?/\\')
    sanitized = ''.join('_' if c in invalid_chars else c for c in title)
    return sanitized[:31]


def add_chart_to_sheet(chart_sheet, source_sheet, chart_title, anchor):
    if source_sheet.max_row < 2:
        return
    chart = BarChart()
    chart.title = chart_title
    chart.style = 2
    chart.x_axis.title = source_sheet.cell(row=1, column=1).value or "Categoría"
    chart.y_axis.title = "Casos"
    data = Reference(source_sheet, min_col=2, min_row=1, max_row=source_sheet.max_row)
    cats = Reference(source_sheet, min_col=1, min_row=2, max_row=source_sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 20
    chart_sheet.add_chart(chart, anchor)


def add_brand_cover_sheet(wb, title, subtitle=None):
    ws = wb.create_sheet(index=0, title="Portada")
    ws["A1"] = APP_NAME
    ws["A2"] = title
    if subtitle:
        ws["A3"] = subtitle
        ws["A3"].font = Font(size=11, bold=True)
    ws["A5"] = "Generado el:"
    ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A1"].font = Font(size=24, bold=True)
    ws["A2"].font = Font(size=14, bold=True)
    ws["A5"].font = Font(size=11, bold=True)
    ws["B5"].font = Font(size=11)
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")
    if subtitle:
        ws.merge_cells("A3:E3")
    ws.column_dimensions[get_column_letter(1)].width = 28
    ws.column_dimensions[get_column_letter(2)].width = 36
    ws.sheet_properties.tabColor = "1e40af"
    return ws


def build_excel_report(df):
    if df.empty:
        return None
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    add_brand_cover_sheet(wb, "Reporte de datos de LCA_PRO")

    datos_sheet = wb.create_sheet("Datos")
    datos_sheet.sheet_properties.tabColor = "2563eb"
    sanitized_df = df.copy()
    for col in sanitized_df.columns:
        sanitized_df[col] = sanitized_df[col].apply(normalize_excel_value)
    write_df_to_sheet(datos_sheet, sanitized_df)

    stats_df = build_report_stats(df)
    stats_sheet = wb.create_sheet("Estadísticas")
    write_df_to_sheet(stats_sheet, stats_df)

    pivots = build_report_pivot_tables(df)
    chart_sheet = wb.create_sheet("Gráficos")
    current_anchor_row = 1
    for sheet_name, pivot_df in pivots.items():
        pivot_sheet = wb.create_sheet(sheet_name)
        write_df_to_sheet(pivot_sheet, pivot_df)
        add_chart_to_sheet(chart_sheet, pivot_sheet, f"Gráfico {sheet_name}", f"A{current_anchor_row}")
        current_anchor_row += 20

    if not pivots:
        chart_sheet.append(["No hay datos suficientes para gráficos."])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_case_excel(case_id):
    case_data = get_case(case_id)
    if not case_data:
        return None
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    add_brand_cover_sheet(wb, f"Caso {case_data['id']}", subtitle=case_data.get('unidad', ''))

    summary = pd.DataFrame([{
        "ID Caso": case_data['id'],
        "Empresa": get_empresa_nombre(case_data['empresa_id']),
        "Unidad": case_data['unidad'],
        "Placa": case_data['placa'],
        "Componente": case_data['componente'],
        "Motivo": case_data['motivo'],
        "Horas": case_data['horas'],
        "Observaciones": case_data['observaciones'],
        "Diagnóstico Base": case_data['diagnostico_base'],
        "Creado": case_data['created_at'],
    }])
    summary_sheet = wb.create_sheet("Resumen Caso")
    summary_sheet.sheet_properties.tabColor = "2563eb"
    write_df_to_sheet(summary_sheet, summary)

    for area in AREAS:
        doc = get_doc(case_id, area)
        area_df = pd.DataFrame([doc]) if doc else pd.DataFrame()
        area_title = sanitize_sheet_title(AREA_LABELS[area])
        area_sheet = wb.create_sheet(area_title)
        if area_df.empty:
            area_sheet.append(["No hay datos disponibles para esta área."])
        else:
            write_df_to_sheet(area_sheet, area_df)

    approvals = q("SELECT area, estado, responsable, comentario, fecha FROM approvals WHERE case_id=? ORDER BY id", (case_id,), fetch=True)
    approvals_df = pd.DataFrame(approvals)
    approvals_sheet = wb.create_sheet("Aprobaciones")
    if approvals_df.empty:
        approvals_sheet.append(["No hay aprobaciones registradas."])
    else:
        write_df_to_sheet(approvals_sheet, approvals_df)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_report_pdf(df, title):
    if not REPORTLAB_OK or df.empty:
        return None
    buffer = io.BytesIO()
    page_size = landscape(A4) if len(df.columns) > 7 else A4
    pdf = SimpleDocTemplate(buffer, pagesize=page_size, leftMargin=1.2*cm, rightMargin=1.2*cm, topMargin=1.7*cm, bottomMargin=1.2*cm)
    styles = build_pdf_styles()

    def header_footer(canvas, doc_obj):
        canvas.saveState()
        draw_pdf_watermark(canvas, page_size)
        draw_pdf_logo(canvas, page_size)
        canvas.setFont("Helvetica-Bold", 9)
        canvas.drawString(1.2*cm + 90, page_size[1] - 1.0*cm, APP_NAME)
        canvas.setFont("Helvetica", 8)
        canvas.drawRightString(page_size[0] - 1.2*cm, page_size[1] - 1.0*cm, title)
        canvas.line(1.2*cm, page_size[1] - 1.25*cm, page_size[0] - 1.2*cm, page_size[1] - 1.25*cm)
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(page_size[0] - 1.2*cm, 1.0*cm, f"Página {doc_obj.page}")
        canvas.restoreState()

    story = [
        Paragraph(APP_NAME, styles["CorpTitle"]),
        Paragraph(title, styles["CorpSub"]),
        Spacer(1, 0.2*cm),
    ]

    stats_df = build_report_stats(df)
    story.append(Paragraph("Resumen Ejecutivo", styles["CorpSub"]))
    story.append(build_wrapped_table([stats_df.columns.tolist()] + stats_df.fillna("").astype(str).values.tolist(), header_bg="#0b1736"))
    story.append(Spacer(1, 0.2*cm))

    data = [list(df.columns)] + df.fillna("").astype(str).values.tolist()
    col_count = len(df.columns)
    col_width = (page_size[0] - 2.4 * cm) / max(col_count, 1)
    table = Table(data, colWidths=[col_width] * col_count, hAlign='LEFT', repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d1d5db")),
    ]))
    story.append(table)
    pdf.build(story, onFirstPage=header_footer, onLaterPages=header_footer)
    buffer.seek(0)
    return buffer.read()


def build_wrapped_table(data, col_widths=None, header_bg="#0b1736"):
    styles = build_pdf_styles()
    wrapped = []
    for r_idx, row in enumerate(data):
        new_row = []
        for cell in row:
            txt = "" if cell is None else str(cell)
            style = styles["CorpSub"] if r_idx == 0 else styles["CorpBody"]
            new_row.append(Paragraph(txt.replace("\n", "<br/>"), style))
        wrapped.append(new_row)
    table = Table(wrapped, colWidths=col_widths, hAlign='LEFT', repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor(header_bg)),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#94a3b8")),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
        ("RIGHTPADDING", (0,0), (-1,-1), 5),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    return table

def build_area_pdf(case_id, area):
    if not REPORTLAB_OK:
        return b""
    styles = build_pdf_styles()
    case_data = get_case(case_id)
    doc = get_doc(case_id, area)
    approval = get_approval(case_id, area) or {}
    buffer = io.BytesIO()
    page_size = landscape(A4) if area == "seguridad_sst" else A4

    def header_footer(canvas, doc_obj):
        canvas.saveState()
        draw_pdf_watermark(canvas, page_size)
        draw_pdf_logo(canvas, page_size)
        canvas.setFont("Helvetica-Bold", 9)
        canvas.drawString(1.2*cm + 90, page_size[1]-1.0*cm, APP_NAME)
        canvas.setFont("Helvetica", 8)
        canvas.drawRightString(page_size[0]-1.2*cm, page_size[1]-1.0*cm, PDF_TITLES[area])
        canvas.line(1.2*cm, page_size[1]-1.25*cm, page_size[0]-1.2*cm, page_size[1]-1.25*cm)
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(page_size[0]-1.2*cm, 1.0*cm, f"Página {doc_obj.page}")
        canvas.restoreState()

    pdf = SimpleDocTemplate(buffer, pagesize=page_size, leftMargin=1.2*cm, rightMargin=1.2*cm, topMargin=1.7*cm, bottomMargin=1.2*cm)
    story = [
        Paragraph(APP_NAME, styles["CorpTitle"]),
        Paragraph(PDF_TITLES[area], styles["CorpSub"]),
        Spacer(1, 0.15*cm),
        build_wrapped_table([
            ["Unidad", "Placa", "Componente", "Motivo", "Horas", "Modelo"],
            [case_data.get("unidad",""), case_data.get("placa",""), case_data.get("componente",""), case_data.get("motivo",""), str(case_data.get("horas","")), case_data.get("ai_model","")]
        ], col_widths=[3.1*cm, 2.4*cm, 3.2*cm, 2.8*cm, 1.5*cm, 4.2*cm]),
        Spacer(1, 0.2*cm),
    ]

    if area == "logistica_compras":
        rows = [["Item", "Descripcion", "Cantidad", "Sub Total"]]
        rows.append(["1", doc.get("piezas","") or "Repuesto principal", "1", str(doc.get("costo_piezas",""))])
        rows.append(["2", doc.get("consumibles","") or "Consumibles", "1", ""])
        rows.append(["3", doc.get("mano_obra","") or "Mano de obra", "1", str(doc.get("costo_mo",""))])
        rows.append(["", "", "Subtotal general", str(doc.get("total",""))])
        rows.append(["", "", "Impuestos", ""])
        rows.append(["", "", "Total general con impuestos", ""])
        story.append(build_wrapped_table(rows, col_widths=[1.0*cm, 11.8*cm, 2.2*cm, 3.2*cm]))
        story.append(Spacer(1, 0.18*cm))
        story.append(Paragraph("Términos y condiciones", styles["CorpSub"]))
        story.append(Paragraph(doc.get("justificacion","") or "Entrega sujeta a aprobación, disponibilidad y validación técnica de compra.", styles["CorpBody"]))
    elif area == "seguridad_sst":
        story.append(Paragraph("KPI de seguridad", styles["CorpSub"]))
        story.append(build_wrapped_table([
            ["Indicador", "Formula / Base", "Valor", "Meta", "Estado"],
            ["Cumplimiento inspeccion", "Inspecciones planificadas", "95%", "100%", "En seguimiento"],
            ["Hallazgos criticos", "Total hallazgos", "1", "0", "Accion inmediata"],
            ["Tiempo respuesta", "Horas desde reporte", "6h", "<4h", "Mejorar"],
        ], col_widths=[5.0*cm, 7.2*cm, 2.0*cm, 2.0*cm, 3.5*cm], header_bg="#f59e0b"))
        story.append(Spacer(1, 0.22*cm))
        story.append(Paragraph("Matriz IPRC", styles["CorpSub"]))
        story.append(build_wrapped_table([
            ["Peligro", "Riesgo", "Nivel", "Control", "Responsable", "Plazo"],
            [doc.get("peligro",""), doc.get("riesgo",""), "Alto", doc.get("controles",""), "Mantenimiento", "24h"],
            ["Fuga", "Contaminacion", "Medio", doc.get("ats",""), "Tecnica", "48h"],
        ], col_widths=[4.6*cm, 4.8*cm, 1.6*cm, 8.0*cm, 2.8*cm, 1.8*cm], header_bg="#0ea5e9"))
    else:
        data = [["Campo", "Detalle"]]
        for k, v in doc.items():
            data.append([k.replace("_", " ").title(), str(v)])
        story.append(build_wrapped_table(data, col_widths=[5.2*cm, 11.8*cm]))

    story.append(Spacer(1, 0.22*cm))
    story.append(Paragraph("Aprobación", styles["CorpSub"]))
    story.append(build_wrapped_table([
        ["Estado", "Responsable", "Comentario / Sustento", "Fecha"],
        [approval.get("estado","Pendiente"), approval.get("responsable",""), approval.get("comentario",""), approval.get("fecha","")]
    ], col_widths=[2.8*cm, 3.5*cm, 10.0*cm, 2.5*cm]))
    pdf.build(story, onFirstPage=header_footer, onLaterPages=header_footer)
    return buffer.getvalue()

def build_package_pdf(case_id):
    if not REPORTLAB_OK:
        return b""
    styles = build_pdf_styles()
    case_data = get_case(case_id)
    buffer = io.BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=1.2*cm, rightMargin=1.2*cm, topMargin=1.7*cm, bottomMargin=1.2*cm)
    def header_footer(canvas, doc_obj):
        canvas.saveState()
        draw_pdf_watermark(canvas, A4)
        draw_pdf_logo(canvas, A4)
        canvas.setFont("Helvetica-Bold", 9)
        canvas.drawString(1.2*cm + 90, A4[1]-1.0*cm, APP_NAME)
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(A4[0] - 1.2*cm, 1.0*cm, f"Página {doc_obj.page}")
        canvas.restoreState()
    story = [
        Paragraph(APP_NAME, styles["CorpTitle"]),
        Paragraph("Paquete consolidado de documentos", styles["CorpSub"]),
        Spacer(1, 0.15*cm),
        build_wrapped_table([
            ["Unidad", "Placa", "Componente", "Motivo", "Horas", "Estado Global"],
            [case_data.get("unidad",""), case_data.get("placa",""), case_data.get("componente",""), case_data.get("motivo",""), str(case_data.get("horas","")), compute_global_status(case_id)]
        ], col_widths=[3.1*cm, 2.4*cm, 3.2*cm, 2.8*cm, 1.5*cm, 4.2*cm]),
        Spacer(1, 0.2*cm),
    ]
    for idx, area in enumerate(AREAS):
        doc = get_doc(case_id, area)
        ap = get_approval(case_id, area) or {}
        story.append(Paragraph(PDF_TITLES[area], styles["CorpSub"]))
        rows = [["Campo", "Detalle"]]
        for k, v in doc.items():
            rows.append([k.replace("_", " ").title(), str(v)])
        rows.append(["Estado", ap.get("estado","Pendiente")])
        rows.append(["Responsable", ap.get("responsable","")])
        rows.append(["Comentario", ap.get("comentario","")])
        story.append(build_wrapped_table(rows, col_widths=[5.0*cm, 12.0*cm]))
        if idx < len(AREAS)-1:
            story.append(PageBreak())
    pdf.build(story, onFirstPage=header_footer, onLaterPages=header_footer)
    return buffer.getvalue()

def approval_widget(case_id, area):
    st.markdown(f'<div class="card"><div class="section-title">Aprobación - {AREA_LABELS[area]}</div>', unsafe_allow_html=True)
    ap = get_approval(case_id, area) or {}
    with st.form(f"approval_form_{area}_{case_id}"):
        c1, c2 = st.columns([1, 1.2])
        with c1:
            responsable = st.text_input(f"Responsable {AREA_LABELS[area]}", value=ap.get("responsable",""))
            estado = st.selectbox(f"Estado {AREA_LABELS[area]}", ESTADOS, index=ESTADOS.index(ap.get("estado","Pendiente")))
        with c2:
            comentario = st.text_area(f"Comentario / Sustento {AREA_LABELS[area]}", value=ap.get("comentario",""), height=120)
        if st.form_submit_button(f"Guardar aprobación - {AREA_LABELS[area]}", use_container_width=True):
            if estado == "Desaprobado" and not comentario.strip():
                st.error("Si desapruebas, el comentario es obligatorio.")
            elif not responsable.strip():
                st.error("El responsable es obligatorio.")
            else:
                save_approval(case_id, area, estado, responsable.strip(), comentario.strip())
                st.success("Aprobación guardada.")
    latest = get_approval(case_id, area)
    if latest:
        st.info(f"Estado actual: {latest['estado']} | Responsable: {latest['responsable']} | Fecha: {latest['fecha']}")
    st.markdown("</div>", unsafe_allow_html=True)

def area_editor(case_id, area, fields, title):
    doc = get_doc(case_id, area)
    st.markdown(f'<div class="card"><div class="section-title">{title}</div>', unsafe_allow_html=True)
    with st.form(f"form_{area}_{case_id}"):
        values = {}
        for f in fields:
            if f.get("type") == "textarea":
                values[f["key"]] = st.text_area(f["label"], value=doc.get(f["key"], ""), height=f.get("height", 120))
            elif f.get("type") == "select":
                options = f["options"]
                current = doc.get(f["key"], options[0])
                idx = options.index(current) if current in options else 0
                values[f["key"]] = st.selectbox(f["label"], options, index=idx)
            else:
                values[f["key"]] = st.text_input(f["label"], value=str(doc.get(f["key"], "")))
        if st.form_submit_button(f"Guardar {title}", use_container_width=True):
            if area == "logistica_compras":
                try:
                    values["total"] = str(float(values.get("costo_piezas", "0") or 0) + float(values.get("costo_mo", "0") or 0))
                except Exception:
                    values["total"] = ""
            save_doc(case_id, area, values)
            st.success(f"{title} guardado.")
    current = get_doc(case_id, area)
    if EXCEL_OK:
        excel_bytes = build_case_excel(case_id)
        if excel_bytes:
            st.download_button(
                label="📥 Descargar paquete Excel consolidado",
                data=excel_bytes,
                file_name=f"{APP_NAME}_paquete_caso_{case_id}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"excel_paquete_caso_{case_id}_{area}"
            )
        else:
            st.warning("No se pudo generar el paquete Excel del caso")
    else:
        st.warning("OpenPyXL no disponible, exportación a Excel deshabilitada")
    if current:
        st.dataframe(pd.DataFrame([current]), use_container_width=True)
    if REPORTLAB_OK:
        st.download_button(
            label=f"📄 Descargar PDF - {title}",
            data=build_area_pdf(case_id, area),
            file_name=f"{APP_NAME}_{area}_{case_id}.pdf",
            mime="application/pdf",
            use_container_width=True,
            key=f"pdf_area_{area}_{case_id}"
        )
    st.markdown("</div>", unsafe_allow_html=True)
    approval_widget(case_id, area)

def trazabilidad_tab(case_id):
    st.markdown('<div class="card"><div class="section-title">Trazabilidad y PDFs</div>', unsafe_allow_html=True)
    rows = q("SELECT area, estado, responsable, comentario, fecha FROM approvals WHERE case_id=? ORDER BY id DESC", (case_id,), fetch=True)
    latest = {}
    for r in rows:
        if r["area"] not in latest:
            latest[r["area"]] = r
    table = []
    for area in AREAS:
        r = latest.get(area, {"estado":"Pendiente","responsable":"","comentario":"","fecha":""})
        table.append({"Área": AREA_LABELS[area], "Estado": r["estado"], "Responsable": r["responsable"], "Fecha": r["fecha"], "Comentario": r["comentario"]})
    st.dataframe(pd.DataFrame(table), use_container_width=True)
    render_status_badge(compute_global_status(case_id))
    if st.session_state.get("empresa_nombre") == "EQUIPRO Perú" or st.session_state.usuario_rol == "admin":
        if REPORTLAB_OK:
            st.download_button(
                label="📦 Descargar paquete PDF consolidado",
                data=build_package_pdf(case_id),
                file_name=f"{APP_NAME}_paquete_caso_{case_id}.pdf",
                mime="application/pdf",
                use_container_width=True,
                key=f"pdf_paquete_general_{case_id}"
            )
        if EXCEL_OK:
            excel_bytes = build_case_excel(case_id)
            if excel_bytes:
                st.download_button(
                    label="📥 Descargar paquete Excel consolidado",
                    data=excel_bytes,
                    file_name=f"{APP_NAME}_paquete_caso_{case_id}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"excel_paquete_general_{case_id}"
                )
            else:
                st.warning("No se pudo generar el paquete Excel del caso")
        else:
            st.warning("OpenPyXL no disponible, exportación a Excel deshabilitada")
    st.markdown("</div>", unsafe_allow_html=True)

def expert_tab(case_id, case_data):
    st.markdown('<div class="card"><div class="section-title">Motor experto / Evidencias</div><div class="small-note">Versión corta y estable.</div>', unsafe_allow_html=True)
    with st.form(f"evidencias_form_{case_id}"):
        image_file = st.file_uploader("Adjuntar Imagen", type=["png","jpg","jpeg","bmp","webp"], key=f"img_{case_id}")
        audio_file = st.file_uploader("Adjuntar Audio", type=["mp3","wav","m4a","ogg","flac"], key=f"aud_{case_id}")
        video_file = st.file_uploader("Adjuntar Video", type=["mp4","avi","mov","mkv","wmv","webm"], key=f"vid_{case_id}")
        c1, c2, c3 = st.columns(3)
        with c1:
            sensor_tipo = st.text_input("Tipo sensor", value=case_data.get("sensor_tipo",""))
            sensor_valor = st.text_input("Valor sensor", value=case_data.get("sensor_valor",""))
        with c2:
            sensor_unidad = st.text_input("Unidad sensor", value=case_data.get("sensor_unidad",""))
            sensor_min = st.text_input("Rango mínimo", value=case_data.get("sensor_min",""))
        with c3:
            sensor_max = st.text_input("Rango máximo", value=case_data.get("sensor_max",""))
            sensor_obs = st.text_input("Observación sensor", value=case_data.get("sensor_obs",""))
        if st.form_submit_button("💾 Guardar evidencias", use_container_width=True):
            img_path = save_upload(image_file, f"case{case_id}_img") if image_file else case_data.get("image_path","")
            aud_path = save_upload(audio_file, f"case{case_id}_aud") if audio_file else case_data.get("audio_path","")
            vid_path = save_upload(video_file, f"case{case_id}_vid") if video_file else case_data.get("video_path","")
            q("""UPDATE cases SET image_path=?, audio_path=?, video_path=?, sensor_tipo=?, sensor_valor=?, sensor_unidad=?, sensor_min=?, sensor_max=?, sensor_obs=? WHERE id=?""",
              (img_path, aud_path, vid_path, sensor_tipo, sensor_valor, sensor_unidad, sensor_min, sensor_max, sensor_obs, case_id))
            st.success("Evidencias guardadas.")
    current_case = get_case(case_id)
    if current_case.get("image_path"):
        try:
            st.image(current_case["image_path"], caption="Imagen guardada", use_container_width=True)
        except Exception:
            pass
    if st.button("🚀 Ejecutar motor experto", use_container_width=True):
        image_info = analyze_image(current_case.get("image_path",""))
        audio_info = analyze_audio(current_case.get("audio_path",""))
        video_info = analyze_video(current_case.get("video_path",""))
        data = {
            "unidad": current_case["unidad"], "placa": current_case["placa"], "componente": current_case["componente"],
            "motivo": current_case["motivo"], "horas": current_case["horas"], "obs": current_case["observaciones"],
            "sensor_tipo": current_case.get("sensor_tipo",""), "sensor_valor": current_case.get("sensor_valor",""),
            "sensor_unidad": current_case.get("sensor_unidad",""), "sensor_min": current_case.get("sensor_min",""),
            "sensor_max": current_case.get("sensor_max",""), "sensor_obs": current_case.get("sensor_obs",""),
        }
        prompt_key, system_prompt, user_prompt = build_expert_prompts(data, image_info, audio_info, video_info)
        with st.spinner("Ejecutando motor experto..."):
            result, model, errors = call_ai(system_prompt, user_prompt)
        cols = st.columns(3)
        cols[0].write("Imagen"); cols[0].json(image_info)
        cols[1].write("Audio"); cols[1].json(audio_info)
        cols[2].write("Video"); cols[2].json(video_info)
        if not result:
            st.error("No fue posible generar resultado experto.")
            st.code("\n".join(errors) if errors else "Sin errores")
        else:
            parsed = parse_ai_json(result)
            if not parsed:
                st.error("La respuesta del motor no devolvió JSON válido.")
                with st.expander("Respuesta cruda"):
                    st.code(result)
            else:
                q("UPDATE cases SET diagnostico_base=?, ai_model=?, prompt_key=? WHERE id=?", (parsed.get("diagnostico_base",""), model or "", prompt_key, case_id))
                for area in AREAS:
                    save_doc(case_id, area, parsed.get(area, {}))
                st.success(f"Motor experto ejecutado con modelo: {model}")
                with st.expander("Ver JSON generado"):
                    st.code(json.dumps(parsed, ensure_ascii=False, indent=2))
    st.markdown("</div>", unsafe_allow_html=True)


def show_dashboard():
    if not is_report_viewer():
        st.warning("No tiene acceso a este panel. Usa el módulo Casos para ingresar información.")
        return

    st.markdown('<div class="hero"><h1>Dashboard LCA_PRO</h1><p>Resumen general de casos y estado</p></div>', unsafe_allow_html=True)
    
    df_cases = get_cases_df()
    if df_cases.empty:
        st.info("No hay casos aún.")
        return
    
    # KPIs generales
    total_cases = len(df_cases)
    approved = sum(
        1 for case_id in df_cases['id'].astype(int).tolist()
        if compute_global_status(case_id) == "Aprobado Final"
    )
    pending = total_cases - approved
    
    c1, c2, c3 = st.columns(3)
    c1.markdown(f'<div class="kpi kpi-blue"><div class="kpi-title">Total Casos</div><div class="kpi-value">{total_cases}</div></div>', unsafe_allow_html=True)
    c2.markdown(f'<div class="kpi kpi-green"><div class="kpi-title">Aprobados</div><div class="kpi-value">{approved}</div></div>', unsafe_allow_html=True)
    c3.markdown(f'<div class="kpi kpi-amber"><div class="kpi-title">Pendientes</div><div class="kpi-value">{pending}</div></div>', unsafe_allow_html=True)
    
    st.markdown('<div class="card"><div class="section-title">Casos Recientes</div>', unsafe_allow_html=True)
    st.dataframe(df_cases.head(10), use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)

def show_cases():
    st.markdown('<div class="hero"><h1>Casos</h1><p>Gestión de casos técnicos</p></div>', unsafe_allow_html=True)
    
    st.markdown("## Crear caso")
    with st.form("new_case"):
        unidad = st.text_input("Unidad")
        placa = st.text_input("Placa")
        componente = st.text_input("Componente")
        motivo = st.selectbox("Motivo", ["Preventivo", "Correctivo", "Inspección", "Falla reportada"])
        horas = st.number_input("Horas unidad", min_value=0, step=1)
        observaciones = st.text_area("Observaciones", height=100)
        diagnostico_base = st.text_area("Diagnóstico base", height=100)
        if st.form_submit_button("Guardar caso", use_container_width=True):
            case_id = q(
                """INSERT INTO cases(
                    empresa_id, unidad, placa, componente, motivo, horas, observaciones, diagnostico_base,
                    image_path, audio_path, video_path, sensor_tipo, sensor_valor, sensor_unidad,
                    sensor_min, sensor_max, sensor_obs, ai_model, prompt_key, created_at
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (st.session_state.empresa_id, unidad, placa, componente, motivo, horas, observaciones, diagnostico_base, "", "", "", "", "", "", "", "", "", "", "", now_str())
            )
            for area in AREAS:
                q("INSERT INTO area_docs(case_id, area, contenido_json, updated_at) VALUES(?,?,?,?)", (case_id, area, "{}", now_str()))
                q("INSERT INTO approvals(case_id, area, estado, responsable, comentario, fecha) VALUES(?,?,?,?,?,?)", (case_id, area, "Pendiente", "", "", now_str()))
            st.success("Caso creado.")

    df_cases = get_cases_df()
    if df_cases.empty:
        st.info("No hay casos aún.")
        return
    labels = [f"{r['id']} | {r['placa']} | {r['unidad']} | {r['componente']}" for _, r in df_cases.iterrows()]
    selected = st.selectbox("Selecciona un caso", labels)
    case_id = int(selected.split("|")[0].strip())
    case_data = get_case(case_id)
    global_status = compute_global_status(case_id)

    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(f'<div class="kpi kpi-blue"><div class="kpi-title">Unidad</div><div class="kpi-value">{case_data["unidad"]}</div></div>', unsafe_allow_html=True)
    c2.markdown(f'<div class="kpi kpi-green"><div class="kpi-title">Componente</div><div class="kpi-value">{case_data["componente"]}</div></div>', unsafe_allow_html=True)
    color_class = "kpi-green" if global_status == "Aprobado Final" else ("kpi-red" if global_status == "En Observación" else "kpi-amber")
    c4.markdown(f'<div class="kpi {color_class}"><div class="kpi-title">Estado Global</div><div class="kpi-value">{global_status}</div></div>', unsafe_allow_html=True)

    st.markdown('<div class="card"><div class="section-title">Resumen del caso</div>', unsafe_allow_html=True)
    st.dataframe(pd.DataFrame([{
        "Unidad": case_data["unidad"],
        "Placa": case_data["placa"],
        "Componente": case_data["componente"],
        "Motivo": case_data["motivo"],
        "Horas": case_data["horas"],
        "Diagnóstico Base": case_data["diagnostico_base"],
        "Creado": case_data["created_at"],
    }]), use_container_width=True)
    render_status_badge(global_status)
    st.markdown("</div>", unsafe_allow_html=True)

def show_expert():
    st.markdown('<div class="hero"><h1>Motor Experto</h1><p>Análisis técnico con IA</p></div>', unsafe_allow_html=True)
    
    df_cases = get_cases_df()
    if df_cases.empty:
        st.info("No hay casos aún.")
        return
    labels = [f"{r['id']} | {r['placa']} | {r['unidad']} | {r['componente']}" for _, r in df_cases.iterrows()]
    selected = st.selectbox("Selecciona un caso para análisis", labels)
    case_id = int(selected.split("|")[0].strip())
    case_data = get_case(case_id)
    
    expert_tab(case_id, case_data)

def show_case_reports():
    if not is_report_viewer():
        st.warning("No tiene acceso a esta sección. Los técnicos deben usar el módulo Casos para ingresar información.")
        return

    st.markdown('<div class="hero"><h1>Reportes por Caso</h1><p>Documentos y trazabilidad por caso</p></div>', unsafe_allow_html=True)
    
    df_cases = get_cases_df()
    if df_cases.empty:
        st.info("No hay casos aún.")
        return
    labels = [f"{r['id']} | {r['placa']} | {r['unidad']} | {r['componente']}" for _, r in df_cases.iterrows()]
    selected = st.selectbox("Selecciona un caso para reportes", labels)
    case_id = int(selected.split("|")[0].strip())
    
    tabs = st.tabs(["Supervisión Técnica", "Orden de Trabajo", "Seguridad / SST", "Logística / Compras", "Trazabilidad / PDFs"])
    with tabs[0]:
        area_editor(case_id, "supervision_tecnica", [
            {"label":"Hallazgos técnicos","key":"hallazgos","type":"textarea","height":150},
            {"label":"Diagnóstico técnico","key":"diagnostico","type":"textarea","height":150},
            {"label":"Criticidad técnica","key":"criticidad","type":"select","options":["Baja","Media","Alta","Crítica"]},
            {"label":"Pruebas recomendadas","key":"pruebas","type":"textarea","height":120},
            {"label":"Recomendaciones técnicas","key":"recomendaciones","type":"textarea","height":120},
        ], "Supervisión Técnica")
    with tabs[1]:
        area_editor(case_id, "orden_trabajo", [
            {"label":"Actividades a ejecutar","key":"actividades","type":"textarea","height":180},
            {"label":"Personal requerido","key":"personal"},
            {"label":"HH estimadas","key":"hh_estimadas"},
            {"label":"Herramientas / equipos","key":"herramientas","type":"textarea","height":120},
            {"label":"Repuestos preliminares","key":"repuestos","type":"textarea","height":120},
        ], "Orden de Trabajo")
    with tabs[2]:
        area_editor(case_id, "seguridad_sst", [
            {"label":"IPERC - Peligros identificados","key":"peligro","type":"textarea","height":110},
            {"label":"IPERC - Riesgos","key":"riesgo","type":"textarea","height":110},
            {"label":"IPERC - Controles","key":"controles","type":"textarea","height":110},
            {"label":"ATS - Pasos críticos y controles","key":"ats","type":"textarea","height":140},
            {"label":"EPP requerido","key":"epp","type":"textarea","height":90},
            {"label":"KPI de seguridad","key":"kpi"},
        ], "Seguridad / SST")
    with tabs[3]:
        area_editor(case_id, "logistica_compras", [
            {"label":"Piezas / repuestos a comprar","key":"piezas","type":"textarea","height":140},
            {"label":"Consumibles","key":"consumibles","type":"textarea","height":100},
            {"label":"Mano de obra estimada","key":"mano_obra"},
            {"label":"Costo estimado piezas","key":"costo_piezas"},
            {"label":"Costo estimado mano de obra","key":"costo_mo"},
            {"label":"Prioridad de compra","key":"prioridad","type":"select","options":["Normal","Urgente","Programada"]},
            {"label":"Justificación técnica de compra","key":"justificacion","type":"textarea","height":120},
        ], "Logística / Compras")
    with tabs[4]:
        trazabilidad_tab(case_id)


def show_admin():
    if st.session_state.get("usuario_rol") != "admin":
        st.warning("No tiene acceso a esta sección administrativa.")
        return

    st.header("Administración del Sistema")
    
    tab1, tab2 = st.tabs(["Empresas", "Usuarios"])
    
    with tab1:
        st.subheader("Gestión de Empresas")
        
        empresas = get_all_empresas()
        
        for emp in empresas:
            with st.expander(f"{emp['nombre']} ({'Activa' if emp['activa'] else 'Inactiva'})"):
                with st.form(f"edit_empresa_{emp['id']}"):
                    nombre = st.text_input("Nombre", value=emp['nombre'])
                    descripcion = st.text_input("Descripción", value=emp.get('descripcion', ''))
                    email = st.text_input("Email", value=emp.get('email', ''))
                    telefono = st.text_input("Teléfono", value=emp.get('telefono', ''))
                    direccion = st.text_input("Dirección", value=emp.get('direccion', ''))
                    ciudad = st.text_input("Ciudad", value=emp.get('ciudad', ''))
                    pais = st.text_input("País", value=emp.get('pais', ''))
                    activa = st.checkbox("Activa", value=bool(emp['activa']))
                    
                    if st.form_submit_button("Actualizar"):
                        update_empresa(emp['id'], nombre, descripcion, email, telefono, direccion, ciudad, pais, 1 if activa else 0)
                        st.success("Empresa actualizada")
                        st.rerun()
        
        with st.form("new_empresa"):
            st.subheader("Nueva Empresa")
            nombre = st.text_input("Nombre")
            descripcion = st.text_input("Descripción")
            email = st.text_input("Email")
            telefono = st.text_input("Teléfono")
            direccion = st.text_input("Dirección")
            ciudad = st.text_input("Ciudad")
            pais = st.text_input("País")
            
            if st.form_submit_button("Crear Empresa"):
                if nombre:
                    create_empresa(nombre, descripcion, email, telefono, direccion, ciudad, pais)
                    st.success("Empresa creada")
                    st.rerun()
                else:
                    st.error("Nombre requerido")
    
    with tab2:
        st.subheader("Gestión de Usuarios")
        
        empresa_options = {emp['nombre']: emp['id'] for emp in get_empresas()}
        empresa_sel = st.selectbox("Seleccionar Empresa", options=list(empresa_options.keys()))
        empresa_id = empresa_options[empresa_sel]
        
        usuarios = get_usuarios_by_empresa(empresa_id)
        
        for user in usuarios:
            with st.expander(f"{user['usuario']} ({user['rol']}) - {'Activo' if user['activo'] else 'Inactivo'}"):
                with st.form(f"edit_user_{user['id']}"):
                    usuario = st.text_input("Usuario", value=user['usuario'])
                    password = st.text_input("Contraseña", type="password", value=user['password'])
                    nombre_completo = st.text_input("Nombre Completo", value=user.get('nombre_completo', ''))
                    email = st.text_input("Email", value=user.get('email', ''))
                    telefono = st.text_input("Teléfono", value=user.get('telefono', ''))
                    rol = st.selectbox("Rol", ["usuario", "tecnico", "admin"], index=["usuario", "tecnico", "admin"].index(user['rol']))
                    activo = st.checkbox("Activo", value=bool(user['activo']))
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.form_submit_button("Actualizar"):
                            update_usuario(user['id'], usuario, password, nombre_completo, email, telefono, rol, 1 if activo else 0)
                            st.success("Usuario actualizado")
                            st.rerun()
                    with col2:
                        if st.form_submit_button("Eliminar", type="secondary"):
                            delete_usuario(user['id'])
                            st.success("Usuario eliminado")
                            st.rerun()
        
        with st.form("new_user"):
            st.subheader("Nuevo Usuario")
            usuario = st.text_input("Usuario")
            password = st.text_input("Contraseña", type="password")
            nombre_completo = st.text_input("Nombre Completo")
            email = st.text_input("Email")
            telefono = st.text_input("Teléfono")
            rol = st.selectbox("Rol", ["usuario", "tecnico", "admin"])
            
            if st.form_submit_button("Crear Usuario"):
                if usuario and password:
                    create_usuario(empresa_id, usuario, password, nombre_completo, email, telefono, rol)
                    st.success("Usuario creado")
                    st.rerun()
                else:
                    st.error("Usuario y contraseña requeridos")


def show_create_users():
    if st.session_state.get("usuario_rol") != "admin":
        st.warning("No tiene acceso a esta sección administrativa.")
        return

    st.header("Crear Usuarios - EQUIPRO Perú")
    
    st.subheader("Gestión de Usuarios")
    
    empresa_options = {emp['nombre']: emp['id'] for emp in get_empresas()}
    empresa_sel = st.selectbox("Seleccionar Empresa", options=list(empresa_options.keys()))
    empresa_id = empresa_options[empresa_sel]
    
    usuarios = get_usuarios_by_empresa(empresa_id)
    
    for user in usuarios:
        with st.expander(f"{user['usuario']} ({user['rol']}) - {'Activo' if user['activo'] else 'Inactivo'}"):
            with st.form(f"edit_user_{user['id']}"):
                usuario = st.text_input("Usuario", value=user['usuario'])
                password = st.text_input("Contraseña", type="password", value=user['password'])
                nombre_completo = st.text_input("Nombre Completo", value=user.get('nombre_completo', ''))
                email = st.text_input("Email", value=user.get('email', ''))
                telefono = st.text_input("Teléfono", value=user.get('telefono', ''))
                rol = st.selectbox("Rol", ["usuario", "tecnico", "admin"], index=["usuario", "tecnico", "admin"].index(user['rol']))
                activo = st.checkbox("Activo", value=bool(user['activo']))
                
                col1, col2 = st.columns(2)
                with col1:
                    if st.form_submit_button("Actualizar"):
                        update_usuario(user['id'], usuario, password, nombre_completo, email, telefono, rol, 1 if activo else 0)
                        st.success("Usuario actualizado")
                        st.rerun()
                with col2:
                    if st.form_submit_button("Eliminar", type="secondary"):
                        delete_usuario(user['id'])
                        st.success("Usuario eliminado")
                        st.rerun()
    
    with st.form("new_user"):
        st.subheader("Nuevo Usuario")
        usuario = st.text_input("Usuario")
        nombre_completo = st.text_input("Nombre Completo")
        email = st.text_input("Email")
        telefono = st.text_input("Teléfono")
        rol = st.selectbox("Rol", ["usuario", "tecnico", "admin"])
        
        if st.form_submit_button("Crear Usuario"):
            if usuario:
                password = "123456"  # Contraseña por defecto
                create_usuario(empresa_id, usuario, password, nombre_completo, email, telefono, rol)
                st.success(f"Usuario creado. Contraseña por defecto: {password}")
                st.rerun()
            else:
                st.error("Usuario requerido")


def show_global_reports():
    if not is_report_viewer():
        st.warning("No tiene acceso a esta sección. Los técnicos no pueden ver informes globales.")
        return

    st.header("Reportes Globales")
    
    empresa_id = st.session_state.get("empresa_id")
    tabs = st.tabs(["Supervisión Técnica", "Orden de Trabajo", "Seguridad SST", "Logística Compras", "Informe Supervisor"])
    
    for i, area in enumerate(AREAS):
        with tabs[i]:
            st.subheader(f"Reportes - {AREA_LABELS[area]}")
            
            # Get all cases with docs for this area, restricted to la empresa del usuario
            cases = q(
                "SELECT c.id, c.unidad, c.placa, c.componente, c.created_at, a.contenido_json "
                "FROM cases c JOIN area_docs a ON c.id = a.case_id "
                "WHERE a.area = ? AND c.empresa_id = ? ORDER BY c.id",
                (area, empresa_id),
                fetch=True
            )
            
            if cases:
                data = []
                for case in cases:
                    doc = json.loads(case['contenido_json'])
                    row = {
                        "ID Caso": case['id'],
                        "Empresa": st.session_state.get("empresa_nombre", ""),
                        "Unidad": case['unidad'],
                        "Placa": case['placa'],
                        "Componente": case['componente'],
                        "Fecha": case['created_at'],
                    }
                    # Add fields from doc
                    for key, value in doc.items():
                        row[key] = value
                    data.append(row)
                
                df = pd.DataFrame(data)
                st.dataframe(df)
                
                stats_df = build_report_stats(df)
                st.markdown("### Estadísticas generales")
                st.dataframe(stats_df)

                pivots = build_report_pivot_tables(df)
                if "Por Unidad" in pivots:
                    st.markdown("#### Casos por Unidad")
                    st.dataframe(pivots["Por Unidad"])
                if "Por Empresa" in pivots:
                    st.markdown("#### Casos por Empresa")
                    st.dataframe(pivots["Por Empresa"])
                if "Por Componente" in pivots:
                    st.markdown("#### Casos por Componente")
                    st.dataframe(pivots["Por Componente"])

                col1, col2 = st.columns([1, 1])
                excel_bytes = build_excel_report(df) if EXCEL_OK else None
                if EXCEL_OK and excel_bytes:
                    with col1:
                        st.download_button(
                            label="Exportar a Excel",
                            data=excel_bytes,
                            file_name=f"reporte_{area.replace(' ', '_')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key=f"excel_global_{area}"
                        )
                else:
                    with col1:
                        st.warning("OpenPyXL no disponible, exportación a Excel deshabilitada")

                pdf_bytes = build_report_pdf(df, f"Reporte {AREA_LABELS[area]}") if REPORTLAB_OK else None
                if REPORTLAB_OK and pdf_bytes:
                    with col2:
                        st.download_button(
                            label="Exportar a PDF",
                            data=pdf_bytes,
                            file_name=f"reporte_{area.replace(' ', '_')}.pdf",
                            mime="application/pdf",
                            use_container_width=True,
                            key=f"pdf_global_{area}"
                        )
                else:
                    with col2:
                        st.warning("ReportLab no disponible, exportación a PDF deshabilitada")
            else:
                st.info("No hay datos para este área")
    
    with tabs[4]:
        st.subheader("Informe Supervisor")
        
        col1, col2 = st.columns(2)
        with col1:
            date_from = st.date_input("Fecha Inicial", value=datetime.now().date() - timedelta(days=30))
        with col2:
            date_to = st.date_input("Fecha Final", value=datetime.now().date())
        
        if st.button("Generar Informe"):
            # Get cases in date range, restricted a la empresa del usuario
            empresa_id = st.session_state.get("empresa_id")
            cases = q(
                "SELECT id, unidad, placa, componente, created_at FROM cases "
                "WHERE empresa_id = ? AND created_at >= ? AND created_at <= ? ORDER BY created_at",
                (empresa_id, date_from.strftime("%Y-%m-%d"), date_to.strftime("%Y-%m-%d")),
                fetch=True
            )
            
            if cases:
                df = pd.DataFrame(cases)
                df['Fecha'] = pd.to_datetime(df['created_at']).dt.date
                st.write(f"Total de unidades ingresadas: {len(cases)}")
                st.dataframe(df[['id', 'unidad', 'placa', 'componente', 'Fecha']])

                report_df = df[['id', 'unidad', 'placa', 'componente', 'Fecha']]
                report_df = report_df.rename(columns={"id": "ID Caso"})
                stats_df = build_report_stats(report_df)
                st.markdown("### Estadísticas por período")
                st.dataframe(stats_df)

                pivots = build_report_pivot_tables(report_df)
                if "Por Unidad" in pivots:
                    st.markdown("#### Casos por Unidad")
                    st.dataframe(pivots["Por Unidad"])
                if "Por Empresa" in pivots:
                    st.markdown("#### Casos por Empresa")
                    st.dataframe(pivots["Por Empresa"])
                if "Por Componente" in pivots:
                    st.markdown("#### Casos por Componente")
                    st.dataframe(pivots["Por Componente"])

                col1, col2 = st.columns([1, 1])
                excel_bytes = build_excel_report(report_df) if EXCEL_OK else None
                if EXCEL_OK and excel_bytes:
                    with col1:
                        st.download_button(
                            label="Exportar a Excel",
                            data=excel_bytes,
                            file_name="informe_supervisor.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key="excel_supervisor"
                        )
                else:
                    with col1:
                        st.warning("OpenPyXL no disponible, exportación a Excel deshabilitada")

                pdf_bytes = build_report_pdf(report_df, "Informe Supervisor") if REPORTLAB_OK else None
                if REPORTLAB_OK and pdf_bytes:
                    with col2:
                        st.download_button(
                            label="Exportar a PDF",
                            data=pdf_bytes,
                            file_name="informe_supervisor.pdf",
                            mime="application/pdf",
                            use_container_width=True,
                            key="pdf_supervisor"
                        )
                else:
                    with col2:
                        st.warning("ReportLab no disponible, exportación a PDF deshabilitada")
            else:
                st.info("No hay casos en el período seleccionado")


def show_login():
    """Muestra pantalla de login"""
    inject_css()
    
    st.markdown('<div class="hero"><h1>LCA_PRO Login</h1><p>Sistema de autenticacion por empresa</p></div>', unsafe_allow_html=True)
    
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">Iniciar sesion</div>', unsafe_allow_html=True)
        
        empresas = get_empresas()
        if not empresas:
            st.error("No hay empresas disponibles. Por favor contacte al administrador.")
            return False
        
        empresa_options = {emp['nombre']: emp['id'] for emp in empresas}
        empresa_nombre = st.selectbox(
            "Selecciona tu empresa",
            options=list(empresa_options.keys()),
            key="login_empresa"
        )
        
        usuario = st.text_input(
            "",
            key="login_usuario",
            placeholder="Usuario"
        )
        
        password = st.text_input(
            "",
            type="password",
            key="login_password",
            placeholder="Contraseña"
        )
        
        if st.button("Ingresar", use_container_width=True, type="primary"):
            if not usuario or not password:
                st.error("Por favor completa todos los campos.")
            else:
                empresa_id = empresa_options[empresa_nombre]
                user_data = validar_login(usuario, password, empresa_id)
                
                if user_data:
                    st.session_state.logged_in = True
                    st.session_state.usuario_id = user_data['id']
                    st.session_state.usuario_nombre = user_data['usuario']
                    st.session_state.usuario_rol = user_data['rol']
                    st.session_state.empresa_id = empresa_id
                    st.session_state.empresa_nombre = empresa_nombre
                    st.success(f"Bienvenido {usuario}!")
                    return
                else:
                    st.error("Usuario, contrasena o empresa incorrectos.")
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col2:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">Credenciales de Demostracion</div>', unsafe_allow_html=True)
        st.markdown("""
**EQUIPRO Peru**
- usuario1 / 123456
- tecnico1 / 123456

**MINERIA ANDINA**
- usuario2 / 123456
- tecnico2 / 123456

**LOGISTICA EXPRESS**
- usuario3 / 123456
""")
        st.markdown('</div>', unsafe_allow_html=True)
    
    return False


def main():
    st.set_page_config(page_title=APP_NAME, layout="wide")
    st.markdown("", unsafe_allow_html=True)

    if "initialized" not in st.session_state:
        st.session_state.initialized = True

    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False

    init_db()
    migrate_schema()
    init_demo_data()

    if not st.session_state.logged_in:
        show_login()
        return

    inject_css()

    if st.session_state.usuario_rol == "admin":
        menu_options = ["Dashboard", "Casos", "Motor Experto", "Reportes por Caso", "Reportes Globales", "Admin"]
        if st.session_state.get("empresa_nombre") == "EQUIPRO Perú":
            menu_options.append("Crear Usuarios")
    elif st.session_state.usuario_rol == "usuario":
        menu_options = ["Dashboard", "Casos", "Motor Experto", "Reportes por Caso", "Reportes Globales"]
    else:
        menu_options = ["Casos", "Motor Experto"]

    menu = st.sidebar.radio("Navegación", menu_options)

    with st.sidebar:
        st.markdown("---")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown(f"**{st.session_state.usuario_nombre}**")
            st.caption(f"Empresa: {st.session_state.empresa_nombre}")
        with col2:
            if st.button("Salir", use_container_width=True):
                    st.session_state.logged_in = False
                    st.rerun()
    if menu == "Dashboard":
        show_dashboard()
    elif menu == "Casos":
        show_cases()
    elif menu == "Motor Experto":
        show_expert()
    elif menu == "Reportes por Caso":
        show_case_reports()
    elif menu == "Reportes Globales":
        show_global_reports()
    elif menu == "Admin":
        show_admin()
    elif menu == "Crear Usuarios":
        show_create_users()

if __name__ == "__main__":
    main()